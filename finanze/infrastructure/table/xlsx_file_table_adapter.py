from io import BytesIO

from application.ports.file_rw_port import TableRWPort
from domain.exception.exceptions import UnsupportedFileFormat
from domain.export import FileFormat
from domain.file_upload import FileUpload

try:
    import pandas as pd

    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    from openpyxl import Workbook, load_workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class XLSXFileTableAdapter(TableRWPort):
    async def convert(self, rows: list[list[str]], format: FileFormat) -> bytes:
        if format != FileFormat.XLSX:
            raise ValueError("XLSXFileTableAdapter only supports XLSX format")

        if not rows:
            raise ValueError("No rows provided for export")

        if HAS_PANDAS:
            return self._convert_with_pandas(rows)
        elif HAS_OPENPYXL:
            return self._convert_with_openpyxl(rows)
        else:
            raise RuntimeError(
                "Neither pandas nor openpyxl is available for XLSX export"
            )

    def _convert_with_pandas(self, rows: list[list[str]]) -> bytes:
        df = pd.DataFrame(rows)
        output = BytesIO()
        pd.DataFrame(df).to_excel(output, header=False, index=False)
        return output.getvalue()

    def _convert_with_openpyxl(self, rows: list[list[str]]) -> bytes:
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    async def parse(self, upload: FileUpload) -> list[list[str]]:
        if upload.content_type not in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ) and not upload.filename.endswith(".xlsx"):
            raise UnsupportedFileFormat("Unsupported XLSX file format")

        upload.data.seek(0)

        if HAS_PANDAS:
            return self._parse_with_pandas(upload)
        elif HAS_OPENPYXL:
            return self._parse_with_openpyxl(upload)
        else:
            raise RuntimeError(
                "Neither pandas nor openpyxl is available for XLSX parsing"
            )

    def _parse_with_pandas(self, upload: FileUpload) -> list[list[str]]:
        df = pd.read_excel(upload.data, header=None)
        return df.fillna("").astype(str).values.tolist()

    def _parse_with_openpyxl(self, upload: FileUpload) -> list[list[str]]:
        wb = load_workbook(upload.data, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])
        wb.close()
        return rows
